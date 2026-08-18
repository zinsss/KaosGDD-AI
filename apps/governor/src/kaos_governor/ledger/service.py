from __future__ import annotations

from collections import defaultdict
from collections.abc import Callable
from dataclasses import dataclass
import hashlib
import io
import json
import os
import threading
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from kaos_governor.database import connect


CATEGORY_DELTAS = {
    "계좌 수입": (1, 0, 0),
    "계좌 지출": (-1, 0, 0),
    "현금 인출": (-1, 1, 0),
    "계좌 입금": (1, -1, 0),
    "상품권 구입 - 계좌": (-1, 0, 1),
    "현금 수입": (0, 1, 0),
    "현금 지출": (0, -1, 0),
    "상품권 구입 - 현금": (0, -1, 1),
    "상품권 사용": (0, 0, -1),
}
CATEGORIES = tuple(CATEGORY_DELTAS)
MAX_DETAILS_LENGTH = 10_000
MAX_ACTOR_LENGTH = 254
BACKUP_ROOT = Path(os.environ.get("LEDGER_BACKUP_ROOT", "/data/ledger/backups"))
LOCAL_TIMEZONE = ZoneInfo(os.environ.get("TZ", "Asia/Seoul"))

_backup_lock = threading.Lock()
_status_lock = threading.Lock()
_backup_status = {
    "enabled": True,
    "root": str(BACKUP_ROOT),
    "lastSuccessAt": "",
    "lastPath": "",
    "lastError": "",
}


class LedgerConflict(Exception):
    pass


@dataclass(frozen=True)
class LedgerRepository:
    connect: Callable = connect


def local_now() -> datetime:
    return datetime.now(LOCAL_TIMEZONE)


def actor_name(value: object) -> str:
    normalized = str(value or "family").strip()
    return (normalized or "family")[:MAX_ACTOR_LENGTH]


def parse_date(value: object) -> date:
    try:
        return date.fromisoformat(str(value or ""))
    except ValueError as exc:
        raise ValueError("invalid_ledger_date") from exc


def parse_amount(value: object, *, allow_empty: bool = False) -> int | None:
    if value in (None, ""):
        if allow_empty:
            return None
        raise ValueError("invalid_ledger_amount")
    if isinstance(value, bool):
        raise ValueError("invalid_ledger_amount")
    try:
        amount = int(str(value).replace(",", "").strip())
    except (TypeError, ValueError) as exc:
        raise ValueError("invalid_ledger_amount") from exc
    if amount < 0:
        raise ValueError("invalid_ledger_amount")
    return amount


def normalize_details(value: object) -> str:
    normalized = str(value or "").strip()
    if len(normalized) > MAX_DETAILS_LENGTH:
        raise ValueError("ledger_details_too_long")
    return normalized


def deltas_for(category: object, amount: object) -> tuple[str, int, tuple[int, int, int]]:
    normalized = str(category or "").strip()
    if normalized not in CATEGORY_DELTAS:
        raise ValueError("invalid_ledger_category")
    parsed_amount = parse_amount(amount)
    assert parsed_amount is not None
    factors = CATEGORY_DELTAS[normalized]
    return normalized, parsed_amount, tuple(factor * parsed_amount for factor in factors)


def normalize_payload(payload: object) -> dict[str, object]:
    if not isinstance(payload, dict):
        raise ValueError("invalid_ledger_payload")
    category, amount, deltas = deltas_for(payload.get("category"), payload.get("amount"))
    return {
        "entry_date": parse_date(payload.get("date")),
        "category": category,
        "amount": amount,
        "details": normalize_details(payload.get("details")),
        "account_delta": deltas[0],
        "cash_delta": deltas[1],
        "gift_delta": deltas[2],
    }


def _iso(value: object) -> str:
    return value.isoformat() if hasattr(value, "isoformat") else str(value or "")


def _entry_from_row(row) -> dict[str, object]:
    return {
        "id": row[0],
        "sortOrder": int(row[1]),
        "date": _iso(row[2]),
        "category": row[3],
        "amount": int(row[4]) if row[4] is not None else None,
        "details": row[5],
        "accountDelta": int(row[6]),
        "cashDelta": int(row[7]),
        "giftDelta": int(row[8]),
        "sourceRow": row[9],
        "sourceChecksum": row[10],
        "locked": bool(row[11]),
        "revision": int(row[12]),
        "createdBy": row[13],
        "updatedBy": row[14],
        "createdAt": _iso(row[15]),
        "updatedAt": _iso(row[16]),
    }


ENTRY_SELECT = """
SELECT id, sort_order, entry_date, category, amount, details,
       account_delta, cash_delta, gift_delta, source_row, source_checksum,
       locked, revision, created_by, updated_by, created_at, updated_at
FROM family_ledger_entries
WHERE deleted_at IS NULL
ORDER BY sort_order, id
"""


def _with_balances(entries: list[dict[str, object]]) -> tuple[list[dict[str, object]], dict[str, int]]:
    account = cash = gift = 0
    result = []
    for entry in entries:
        account += int(entry["accountDelta"])
        cash += int(entry["cashDelta"])
        gift += int(entry["giftDelta"])
        result.append({**entry, "account": account, "cash": cash, "gift": gift})
    return result, {"account": account, "cash": cash, "gift": gift}


def list_ledger(repository: LedgerRepository | None = None) -> dict[str, object]:
    repo = repository or LedgerRepository()
    with repo.connect() as connection:
        entries = [_entry_from_row(row) for row in connection.execute(ENTRY_SELECT).fetchall()]
    entries, balances = _with_balances(entries)
    return {
        "ok": True,
        "entries": entries,
        "balances": balances,
        "categories": list(CATEGORIES),
        "entryCount": len(entries),
    }


def _audit(connection, entry_id: str, action: str, actor: str, before, after) -> None:
    connection.execute(
        """
        INSERT INTO family_ledger_audit (entry_id, action, actor, before_data, after_data)
        VALUES (%s, %s, %s, %s::jsonb, %s::jsonb)
        """,
        (
            entry_id,
            action,
            actor,
            json.dumps(before, ensure_ascii=False) if before is not None else None,
            json.dumps(after, ensure_ascii=False) if after is not None else None,
        ),
    )


def _row_for_update(connection, entry_id: str) -> dict[str, object]:
    row = connection.execute(
        ENTRY_SELECT.replace("WHERE deleted_at IS NULL", "WHERE id = %s AND deleted_at IS NULL").replace(
            "ORDER BY sort_order, id", "FOR UPDATE"
        ),
        (entry_id,),
    ).fetchone()
    if not row:
        raise ValueError("ledger_entry_not_found")
    return _entry_from_row(row)


def _after_mutation(repository: LedgerRepository | None = None) -> None:
    try:
        write_backup("latest", repository)
    except Exception as exc:
        _set_backup_status(error=type(exc).__name__)


def create_entry(payload: object, actor: object = "family", repository: LedgerRepository | None = None) -> dict[str, object]:
    repo = repository or LedgerRepository()
    normalized = normalize_payload(payload)
    normalized_actor = actor_name(actor)
    entry_id = uuid.uuid4().hex
    with repo.connect() as connection:
        with connection.transaction():
            connection.execute("SELECT pg_advisory_xact_lock(hashtext('family-ledger'))")
            sort_order = connection.execute(
                "SELECT COALESCE(max(sort_order), 0) + 1000 FROM family_ledger_entries"
            ).fetchone()[0]
            row = connection.execute(
                """
                INSERT INTO family_ledger_entries (
                    id, sort_order, entry_date, category, amount, details,
                    account_delta, cash_delta, gift_delta, created_by, updated_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, sort_order, entry_date, category, amount, details,
                          account_delta, cash_delta, gift_delta, source_row, source_checksum,
                          locked, revision, created_by, updated_by, created_at, updated_at
                """,
                (
                    entry_id,
                    sort_order,
                    normalized["entry_date"],
                    normalized["category"],
                    normalized["amount"],
                    normalized["details"],
                    normalized["account_delta"],
                    normalized["cash_delta"],
                    normalized["gift_delta"],
                    normalized_actor,
                    normalized_actor,
                ),
            ).fetchone()
            item = _entry_from_row(row)
            _audit(connection, entry_id, "create", normalized_actor, None, item)
    _after_mutation(repo)
    return list_ledger(repo)


def update_entry(
    entry_id: str,
    payload: object,
    actor: object = "family",
    repository: LedgerRepository | None = None,
) -> dict[str, object]:
    if not re_fullmatch_ledger_id(entry_id):
        raise ValueError("ledger_entry_not_found")
    repo = repository or LedgerRepository()
    normalized = normalize_payload(payload)
    normalized_actor = actor_name(actor)
    if not isinstance(payload, dict):
        raise ValueError("invalid_ledger_payload")
    try:
        base_revision = int(payload.get("baseRevision"))
    except (TypeError, ValueError) as exc:
        raise ValueError("invalid_ledger_revision") from exc
    with repo.connect() as connection:
        with connection.transaction():
            before = _row_for_update(connection, entry_id)
            if before["locked"]:
                raise ValueError("ledger_entry_locked")
            if before["revision"] != base_revision:
                raise LedgerConflict("ledger_revision_conflict")
            row = connection.execute(
                """
                UPDATE family_ledger_entries
                SET entry_date = %s, category = %s, amount = %s, details = %s,
                    account_delta = %s, cash_delta = %s, gift_delta = %s,
                    revision = revision + 1, updated_by = %s, updated_at = now()
                WHERE id = %s
                RETURNING id, sort_order, entry_date, category, amount, details,
                          account_delta, cash_delta, gift_delta, source_row, source_checksum,
                          locked, revision, created_by, updated_by, created_at, updated_at
                """,
                (
                    normalized["entry_date"],
                    normalized["category"],
                    normalized["amount"],
                    normalized["details"],
                    normalized["account_delta"],
                    normalized["cash_delta"],
                    normalized["gift_delta"],
                    normalized_actor,
                    entry_id,
                ),
            ).fetchone()
            after = _entry_from_row(row)
            _audit(connection, entry_id, "update", normalized_actor, before, after)
    _after_mutation(repo)
    return list_ledger(repo)


def delete_entry(
    entry_id: str,
    base_revision: object,
    actor: object = "family",
    repository: LedgerRepository | None = None,
) -> dict[str, object]:
    if not re_fullmatch_ledger_id(entry_id):
        raise ValueError("ledger_entry_not_found")
    repo = repository or LedgerRepository()
    normalized_actor = actor_name(actor)
    try:
        revision = int(base_revision)
    except (TypeError, ValueError) as exc:
        raise ValueError("invalid_ledger_revision") from exc
    with repo.connect() as connection:
        with connection.transaction():
            before = _row_for_update(connection, entry_id)
            if before["locked"]:
                raise ValueError("ledger_entry_locked")
            if before["revision"] != revision:
                raise LedgerConflict("ledger_revision_conflict")
            connection.execute(
                """
                UPDATE family_ledger_entries
                SET deleted_at = now(), revision = revision + 1,
                    updated_by = %s, updated_at = now()
                WHERE id = %s
                """,
                (normalized_actor, entry_id),
            )
            _audit(connection, entry_id, "delete", normalized_actor, before, None)
    _after_mutation(repo)
    return list_ledger(repo)


def re_fullmatch_ledger_id(value: str) -> bool:
    import re

    return bool(re.fullmatch(r"[0-9a-z-]+", str(value or "")))


def _audit_rows(repository: LedgerRepository | None = None):
    repo = repository or LedgerRepository()
    with repo.connect() as connection:
        return connection.execute(
            """
            SELECT id, entry_id, action, actor, before_data, after_data, created_at
            FROM family_ledger_audit
            ORDER BY id
            """
        ).fetchall()


def workbook_bytes(repository: LedgerRepository | None = None) -> bytes:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    ledger = list_ledger(repository)
    workbook = Workbook()
    entries_sheet = workbook.active
    entries_sheet.title = "거래내역"
    headers = ["날짜", "사용 구분", "금액", "상세 내용", "계좌", "현금", "상품권"]
    entries_sheet.append(headers)
    for entry in ledger["entries"]:
        entries_sheet.append(
            [
                entry["date"],
                entry["category"],
                entry["amount"],
                entry["details"],
                entry["account"],
                entry["cash"],
                entry["gift"],
            ]
        )
    entries_sheet.freeze_panes = "A2"
    entries_sheet.auto_filter.ref = f"A1:G{max(1, entries_sheet.max_row)}"
    for index, width in enumerate([14, 22, 14, 48, 16, 16, 16], 1):
        entries_sheet.column_dimensions[chr(64 + index)].width = width
    for cell in entries_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6D597A")
        cell.alignment = Alignment(horizontal="center")
    for row in entries_sheet.iter_rows(min_row=2, min_col=3, max_col=7):
        for cell in row:
            cell.number_format = '#,##0;[Red]-#,##0'

    summary_sheet = workbook.create_sheet("월별요약")
    summary_sheet.append(["월", "계좌 변동", "현금 변동", "상품권 변동", "월말 계좌", "월말 현금", "월말 상품권"])
    monthly = defaultdict(lambda: [0, 0, 0, 0, 0, 0])
    for entry in ledger["entries"]:
        key = str(entry["date"])[:7]
        values = monthly[key]
        values[0] += int(entry["accountDelta"])
        values[1] += int(entry["cashDelta"])
        values[2] += int(entry["giftDelta"])
        values[3:] = [entry["account"], entry["cash"], entry["gift"]]
    for month, values in monthly.items():
        summary_sheet.append([month, *values])
    summary_sheet.freeze_panes = "A2"
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6D597A")
    for row in summary_sheet.iter_rows(min_row=2, min_col=2, max_col=7):
        for cell in row:
            cell.number_format = '#,##0;[Red]-#,##0'
    for column in "ABCDEFG":
        summary_sheet.column_dimensions[column].width = 18

    audit_sheet = workbook.create_sheet("변경기록")
    audit_sheet.append(["번호", "시각", "작업", "사용자", "항목 ID", "변경 전", "변경 후"])
    for row in _audit_rows(repository):
        audit_sheet.append(
            [row[0], _iso(row[6]), row[2], row[3], row[1], json.dumps(row[4], ensure_ascii=False), json.dumps(row[5], ensure_ascii=False)]
        )
    audit_sheet.freeze_panes = "A2"
    for cell in audit_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6D597A")
    for column, width in {"A": 10, "B": 24, "C": 12, "D": 28, "E": 34, "F": 80, "G": 80}.items():
        audit_sheet.column_dimensions[column].width = width

    info_sheet = workbook.create_sheet("정보")
    info_sheet.append(["항목", "값"])
    info_sheet.append(["생성 시각", local_now().isoformat()])
    info_sheet.append(["원본", "KaosGovernor PostgreSQL"])
    info_sheet.append(["거래 수", ledger["entryCount"]])
    info_sheet.append(["계좌 잔액", ledger["balances"]["account"]])
    info_sheet.append(["현금 잔액", ledger["balances"]["cash"]])
    info_sheet.append(["상품권 잔액", ledger["balances"]["gift"]])
    info_sheet["A1"].font = info_sheet["B1"].font = Font(bold=True, color="FFFFFF")
    info_sheet["A1"].fill = info_sheet["B1"].fill = PatternFill("solid", fgColor="6D597A")
    info_sheet.column_dimensions["A"].width = 18
    info_sheet.column_dimensions["B"].width = 48

    stream = io.BytesIO()
    workbook.save(stream)
    data = stream.getvalue()
    load_workbook(io.BytesIO(data), read_only=True).close()
    return data


def _set_backup_status(*, path: str = "", error: str = "") -> None:
    with _status_lock:
        if path:
            _backup_status["lastSuccessAt"] = local_now().isoformat()
            _backup_status["lastPath"] = path
            _backup_status["lastError"] = ""
        elif error:
            _backup_status["lastError"] = error


def backup_status() -> dict[str, object]:
    with _status_lock:
        return dict(_backup_status)


def _atomic_write(path: Path, data: bytes) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_bytes(data)
    os.replace(temporary, path)
    checksum = hashlib.sha256(data).hexdigest()
    path.with_suffix(path.suffix + ".sha256").write_text(f"{checksum}  {path.name}\n", encoding="ascii")
    return checksum


def write_backup(kind: str = "manual", repository: LedgerRepository | None = None) -> dict[str, object]:
    now = local_now()
    with _backup_lock:
        data = workbook_bytes(repository)
        if kind == "latest":
            path = BACKUP_ROOT / "latest.xlsx"
        else:
            path = BACKUP_ROOT / "manual" / f"ledger-{now.strftime('%Y%m%d-%H%M%S')}.xlsx"
        checksum = _atomic_write(path, data)
        _set_backup_status(path=str(path))
        return {"ok": True, "path": str(path), "sha256": checksum, "size": len(data), "createdAt": now.isoformat()}
